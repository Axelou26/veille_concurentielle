"""
Onglet Base de données
Gère la gestion et la maintenance de la base de données
"""

import streamlit as st
import pandas as pd
from datetime import datetime
from database_manager import DatabaseManager
from ao_extractor_v2 import AOExtractorV2
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Code d'autorisation pour les actions de maintenance
MAINTENANCE_CODE = "kristelle123"  # Code à changer selon vos besoins

def verify_maintenance_code(action_name: str = "cette action") -> bool:
    """
    Vérifie si l'utilisateur a entré le bon code de maintenance
    
    Args:
        action_name: Nom de l'action pour personnaliser le message
    
    Returns:
        bool: True si le code est correct, False sinon
    """
    if 'maintenance_authorized' not in st.session_state:
        st.session_state.maintenance_authorized = False
    
    if st.session_state.maintenance_authorized:
        return True
    
    # Afficher un champ pour entrer le code
    st.warning(f"🔒 {action_name.capitalize()} nécessite une autorisation")
    
    col1, col2 = st.columns([2, 1])
    with col1:
        entered_code = st.text_input(
            "Entrez le code de maintenance:",
            type="password",
            key=f"maintenance_code_{action_name.replace(' ', '_')}"
        )
    with col2:
        st.write("")  # Espacement vertical
        st.write("")  # Espacement vertical
        verify_button = st.button("✅ Valider", key=f"verify_button_{action_name.replace(' ', '_')}")
    
    if verify_button and entered_code:
        if entered_code == MAINTENANCE_CODE:
            st.session_state.maintenance_authorized = True
            st.success("✅ Code valide - Accès autorisé")
            st.rerun()
        else:
            st.error("❌ Code incorrect - Accès refusé")
            return False
    elif verify_button and not entered_code:
        st.error("⚠️ Veuillez entrer un code")
    
    return False


def create_formatted_excel(df: pd.DataFrame) -> io.BytesIO:
    """
    Crée un fichier Excel formaté avec tableau Excel depuis un DataFrame
    
    Args:
        df: DataFrame pandas à exporter
        
    Returns:
        BytesIO: Buffer contenant le fichier Excel
    """
    # Créer un workbook et une worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Données Veille"
    
    # Styles pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Style pour les bordures
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Écrire les en-têtes
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Écrire les données
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            
            # Gérer les valeurs NaN
            if pd.isna(value):
                cell.value = None
            else:
                cell.value = value
            
            # Formatage conditionnel selon le type
            if isinstance(value, (int, float)) and not pd.isna(value):
                # Format numérique
                cell.number_format = '#,##0' if isinstance(value, int) else '#,##0.00'
            elif pd.api.types.is_datetime64_any_dtype(type(value)) or isinstance(value, pd.Timestamp):
                # Format date
                cell.number_format = 'DD/MM/YYYY'
            
            # Bordures et alignement
            cell.border = thin_border
            if isinstance(value, (int, float)) and not pd.isna(value):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Ajuster la largeur des colonnes automatiquement
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            try:
                if cell.value:
                    # Calculer la longueur du contenu
                    content_length = len(str(cell.value))
                    if content_length > max_length:
                        max_length = content_length
            except:
                pass
        
        # Définir la largeur avec une marge
        adjusted_width = min(max_length + 2, 50)  # Maximum 50 caractères
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Créer un tableau Excel formaté (Table object)
    if len(df) > 0:
        table_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        table = Table(displayName="TableVeille", ref=table_range)
        
        # Style du tableau (style professionnel bleu)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Ajouter le tableau à la worksheet
        ws.add_table(table)
    
    # Congeler la première ligne (en-tête)
    ws.freeze_panes = "A2"
    
    # Sauvegarder dans un buffer en mémoire
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def render_database_tab(
    db_manager: DatabaseManager,
    ao_extractor: AOExtractorV2 = None
):
    """
    Rend l'onglet Base de données
    
    Args:
        db_manager (DatabaseManager): Gestionnaire de base de données
        ao_extractor (AOExtractorV2, optional): Extracteur d'AO pour les métriques
    """
    st.header("🗄️ Gestion de la Base de Données")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📊 Informations de la Base")
        
        # Statistiques de la base
        stats = db_manager.get_statistics()
        if stats:
            st.write(f"**Total des lots:** {stats.get('total_lots', 0)}")
            st.write(f"**Univers représentés:** {len(stats.get('univers_stats', {}))}")
            st.write(f"**Statuts différents:** {len(stats.get('statut_stats', {}))}")
            st.write(f"**Groupements:** {len(stats.get('groupement_stats', {}))}")
        
        # Métadonnées
        st.subheader("📋 Métadonnées")
        if hasattr(db_manager, 'get_metadata'):
            last_import = db_manager.get_metadata('last_excel_import')
            if last_import:
                st.write(f"**Dernier import:** {last_import.get('import_date', 'N/A')}")
                st.write(f"**Fichier:** {last_import.get('file_path', 'N/A')}")
                st.write(f"**Lignes importées:** {last_import.get('rows_imported', 'N/A')}")
    
    with col2:
        st.subheader("🔧 Actions")
        
        # Export des données
        col_export1, col_export2 = st.columns(2)
        
        with col_export1:
            if st.button("📥 Exporter en Excel (formaté)"):
                try:
                    all_data = db_manager.get_all_data()
                    
                    if all_data.empty:
                        st.warning("⚠️ Aucune donnée à exporter")
                    else:
                        # Créer un fichier Excel formaté en mémoire
                        excel_buffer = create_formatted_excel(all_data)
                        
                        st.download_button(
                            label="💾 Télécharger Excel formaté",
                            data=excel_buffer.getvalue(),
                            file_name=f"export_veille_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        st.success(f"✅ Export Excel prêt : {len(all_data)} lignes")
                        
                except Exception as e:
                    st.error(f"❌ Erreur lors de l'export Excel: {e}")
                    import traceback
                    st.code(traceback.format_exc())
        
        with col_export2:
            if st.button("📄 Exporter en CSV"):
                try:
                    all_data = db_manager.get_all_data()
                    csv_data = all_data.to_csv(index=False)
                    
                    st.download_button(
                        label="💾 Télécharger CSV",
                        data=csv_data,
                        file_name=f"export_veille_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
                    
                    st.success(f"✅ Export CSV prêt : {len(all_data)} lignes")
                    
                except Exception as e:
                    st.error(f"❌ Erreur lors de l'export CSV: {e}")
        
        # Recherche dans la base
        st.subheader("🔍 Recherche")
        search_term = st.text_input("Terme de recherche")
        if search_term:
            if hasattr(db_manager, 'search_data'):
                results = db_manager.search_data(search_term)
                if not results.empty:
                    st.write(f"**Résultats trouvés:** {len(results)}")
                    st.dataframe(results.head(10), width='stretch')
                else:
                    st.info("Aucun résultat trouvé")
        
        # Maintenance de la base
        st.subheader("🔧 Maintenance de la Base")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 Créer sauvegarde"):
                if verify_maintenance_code("créer une sauvegarde"):
                    if hasattr(db_manager, 'create_backup'):
                        if db_manager.create_backup():
                            st.success("✅ Sauvegarde créée!")
                            st.session_state.maintenance_authorized = False  # Réinitialiser après action
                        else:
                            st.error("❌ Erreur création sauvegarde")
                    else:
                        st.warning("⚠️ Fonction de sauvegarde non disponible")
        
        with col2:
            if st.button("🔍 Valider intégrité"):
                if verify_maintenance_code("valider l'intégrité"):
                    if hasattr(db_manager, 'validate_data_integrity'):
                        validation = db_manager.validate_data_integrity()
                        if validation['is_valid']:
                            st.success("✅ Intégrité validée")
                        else:
                            st.warning("⚠️ Problèmes détectés")
                            for issue in validation['issues']:
                                st.write(f"• {issue}")
                        st.session_state.maintenance_authorized = False  # Réinitialiser après action
                    else:
                        st.warning("⚠️ Fonction de validation non disponible")
        
        with col3:
            if st.button("⚡ Optimiser base"):
                if verify_maintenance_code("optimiser la base"):
                    if hasattr(db_manager, 'optimize_database'):
                        if db_manager.optimize_database():
                            st.success("✅ Base optimisée!")
                            st.session_state.maintenance_authorized = False  # Réinitialiser après action
                        else:
                            st.error("❌ Erreur optimisation")
                    else:
                        st.warning("⚠️ Fonction d'optimisation non disponible")
        
        # Informations détaillées de la base
        if hasattr(db_manager, 'get_database_info'):
            with st.expander("📊 Informations détaillées de la base"):
                db_info = db_manager.get_database_info()
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Taille de la base", f"{db_info.get('database_size_mb', 0):.2f} MB")
                    st.metric("Sauvegardes créées", db_info.get('performance_metrics', {}).get('backup_count', 0))
                
                with col2:
                    st.metric("Dernière modification", db_info.get('last_modified', 'N/A')[:10] if db_info.get('last_modified') else 'N/A')
                    st.metric("Dernière sauvegarde", db_info.get('last_backup', 'N/A')[:10] if db_info.get('last_backup') else 'N/A')
                
                # Métriques de performance
                perf_metrics = db_info.get('performance_metrics', {})
                if perf_metrics:
                    st.subheader("📈 Métriques de performance")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Requêtes totales", perf_metrics.get('total_queries', 0))
                    with col2:
                        success_rate = perf_metrics.get('success_rate', 0) * 100
                        st.metric("Taux de succès", f"{success_rate:.1f}%")
                    with col3:
                        avg_time = perf_metrics.get('average_query_time', 0)
                        st.metric("Temps moyen", f"{avg_time:.3f}s")
        
        # Métriques V2 de l'extracteur
        if ao_extractor and hasattr(ao_extractor, 'extraction_metrics'):
            with st.expander("🚀 Métriques Extracteur V2"):
                v2_metrics = ao_extractor.extraction_metrics
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Extractions totales", v2_metrics.get('total_extractions', 0))
                with col2:
                    st.metric("Extractions réussies", v2_metrics.get('successful_extractions', 0))
                with col3:
                    st.metric("Erreurs validation", v2_metrics.get('validation_errors', 0))
                with col4:
                    st.metric("Patterns améliorés", v2_metrics.get('pattern_improvements', 0))
        
        # Nettoyage de la base
        if st.button("🗑️ Vider la base de données", type="secondary"):
            if verify_maintenance_code("vider la base de données"):
                st.warning("⚠️ ATTENTION : Cette action est irréversible !")
                if st.checkbox("Confirmer la suppression de toutes les données"):
                    try:
                        cursor = db_manager.connection.cursor()
                        cursor.execute("DELETE FROM appels_offres")
                        db_manager.connection.commit()
                        st.success("✅ Base de données vidée")
                        st.session_state.maintenance_authorized = False  # Réinitialiser après action
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Erreur lors du nettoyage: {e}")



